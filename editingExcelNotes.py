# Module 14 - Excel, Word and PDF Documents: Editing Excel Spreadsheets

import openpyxl
wb = openpyxl.Workbook()

wb.get_sheet_names() # Returns ['Sheet']
sheet = wb.get_sheet_by_name('Sheet')

sheet # Returns <Workbook "Sheet">

sheet['A1'].value # Returns nothing as the sheet is blank.
sheet['A1'].value == None # Returns True.

# To add/edit the sheet treat each cell as an assignment variable

# Example:
sheet['A1'] = 42
# sheet['A1'].value now returns 42 as an integer.
sheet['A2'] = 'Hello'

### The workbook currently only exists in the computer's memory ###
import os
os.chdir('/Users/paulmackay/Desktop/Python/Excel, Word and PDF Documents')
wb.save('example2.xlsx') # Saves the file

### Be sure to save each iteration of the file under a new name ###
### This makes it easier to use older versions should bugs occur ###

sheet2 = wb.create_sheet()
wb.get_sheet_names() # Returns ['Sheet', 'Sheet1']

sheet2.title = 'My New Sheet Name' # This changes the name of the sheet
wb.get_sheet_names() # Now returns ['Sheet', 'My New Sheet Name']

wb.save('example3.xlsx')

wb.create_sheet(index=0, title='My Other Sheet') # Adds new sheet to the front.
wb.save('example4.xlsx')
