# Module 14 - Excel, Word and PDF Documents: Reading Excel Spreadsheets

import openpyxl
import os
os.chdir('/Users/paulmackay/Desktop/Python/Excel, Word and PDF Documents')

workbook = openpyxl.load_workbook('example.xlsx')
type(workbook) # Returns <class 'openpyxl.workbook.workbook.Workbook'>

sheet = workbook.get_sheet_by_name('Sheet1')

# If you do not know the name of the sheet.
workbook.get_sheet_names() # Returns ['Sheet1', 'Sheet2', 'Sheet3']

sheet['A1'] # Returns <Cell 'Sheet1'.A1> which is a cell object.

cell = sheet['A1']
cell.value # Returns datetime.datetime(2015, 4, 5, 13, 34, 2)
# which are the actual contents of the cell.

# To get the string value of the cell.
str(sheet['A1'].value) # Returns '2015-04-05 13:34:02'

sheet['B1'].value # Returns 'Apples'
sheet['C1'].value # Returns 73 (as an integer object).
str(sheet['C1'].value) # Returns the string value '73'.

sheet.cell(row=1, column=2) # Returns <Cell Sheet1.B1>
### This is the same as using "sheet['B1']" ###

# Example using sheet.cell: This allows us to use FOR LOOPS easily
for i in range(1,8):
    print(i, sheet.cell(row=i, column=2).value)

# Returns:
# 1 Apples
# 2 Cherries
# 3 Pears
# 4 Oranges
# 5 Apples
# 6 Bananas
# 7 Strawberries
